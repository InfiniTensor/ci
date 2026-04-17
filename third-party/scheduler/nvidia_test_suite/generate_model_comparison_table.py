import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_model_comparison_excel(output_file, engine=None):
    """
    生成模型性能表格（仅生成表头，不预填充模型）
    
    Args:
        output_file: 输出文件路径
        engine: 推理引擎，可选值：
                None - 生成对比表格（SigInfer vs VLLM）
                "SigInfer" - 仅生成SigInfer报告
                "VLLM_0.9.1" - 仅生成VLLM报告
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # 样式定义
    light_blue_fill = PatternFill("solid", fgColor="9BC2E6")  # 浅蓝色 - 引擎标题
    yellow_fill = PatternFill("solid", fgColor="FFD966")       # 黄色 - 子表头
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    if engine is None:
        # 对比模式：生成SigInfer vs VLLM的对比表格
        ws.title = "Model Comparison"
        
        # 第1行：顶部表头
        # 列1: 模型
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        cell = ws.cell(1, 1, value="模型")
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border

        # 列2-3: SigInfer
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        cell = ws.cell(1, 2, value="SigInfer")
        cell.fill = light_blue_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).fill = light_blue_fill

        # 列4-5: VLLM 0.9.1
        ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=5)
        cell = ws.cell(1, 4, value="VLLM 0.9.1")
        cell.fill = light_blue_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(1, 5).border = thin_border
        ws.cell(1, 5).fill = light_blue_fill

        # 第2行：子表头
        sub_headers = ["opencompass", "SGLang", "opencompass", "SGLang"]
        for idx, header in enumerate(sub_headers, start=2):
            cell = ws.cell(2, idx, value=header)
            cell.fill = yellow_fill
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.border = thin_border

        # 设置列宽
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 30
            
    else:
        # 单引擎模式：只生成一个引擎的报告
        ws.title = f"{engine} Report"
        
        # 第1行：顶部表头
        # 列1: 模型
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        cell = ws.cell(1, 1, value="模型")
        cell.fill = yellow_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border

        # 列2-3: 引擎名称
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=3)
        cell = ws.cell(1, 2, value=engine)
        cell.fill = light_blue_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
        ws.cell(1, 3).border = thin_border
        ws.cell(1, 3).fill = light_blue_fill

        # 第2行：子表头
        sub_headers = ["opencompass", "SGLang"]
        for idx, header in enumerate(sub_headers, start=2):
            cell = ws.cell(2, idx, value=header)
            cell.fill = yellow_fill
            cell.font = Font(bold=True)
            cell.alignment = center_align
            cell.border = thin_border

        # 设置列宽
        ws.column_dimensions['A'].width = 35
        for col in ['B', 'C']:
            ws.column_dimensions[col].width = 30

    # 冻结表头
    ws.freeze_panes = "A3"

    # 设置行高
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 25

    wb.save(output_file)
    print(f"Excel 文件已生成: {output_file}")


def fill_model_comparison_data(excel_file, data_dict, engine=None):
    """
    动态填充模型数据到Excel表格
    
    Args:
        excel_file: Excel文件路径
        data_dict: 数据字典，格式为:
                  {
                      "模型名称": {
                          "SigInfer_opencompass": "数据内容",
                          "SigInfer_SGLang": "数据内容",
                          "VLLM_opencompass": "数据内容",
                          "VLLM_SGLang": "数据内容"
                      }
                  }
        engine: 推理引擎，可选值：
                None - 对比模式（填充所有列）
                "SigInfer" - 仅填充SigInfer数据
                "VLLM_0.9.1" - 仅填充VLLM数据
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 样式定义
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 根据engine参数确定列索引映射
    if engine is None:
        # 对比模式：所有列
        column_mapping = {
            "SigInfer_opencompass": 2,
            "SigInfer_SGLang": 3,
            "VLLM_opencompass": 4,
            "VLLM_SGLang": 5
        }
    elif engine == "SigInfer":
        # 仅SigInfer列
        column_mapping = {
            "SigInfer_opencompass": 2,
            "SigInfer_SGLang": 3
        }
    elif engine == "VLLM_0.9.1":
        # 仅VLLM列
        column_mapping = {
            "VLLM_opencompass": 2,
            "VLLM_SGLang": 3
        }
    else:
        raise ValueError(f"不支持的引擎类型: {engine}，支持的值为: None, 'SigInfer', 'VLLM_0.9.1'")

    # 从第3行开始动态添加数据行
    row_idx = 3
    
    # 遍历数据字典，为每个模型创建一行
    for model_name, model_data in data_dict.items():
        # 第1列：模型名称
        cell = ws.cell(row_idx, 1, value=model_name)
        cell.alignment = left_align
        cell.border = thin_border
        cell.font = Font(bold=True)
        
        # 填充数据列
        for data_key, col_idx in column_mapping.items():
            if data_key in model_data:
                cell = ws.cell(row_idx, col_idx, value=model_data[data_key])
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.border = thin_border
            else:
                # 如果没有数据，也要添加空单元格以保持格式
                cell = ws.cell(row_idx, col_idx, value="")
                cell.alignment = left_align
                cell.border = thin_border
        
        # 设置数据行高度
        ws.row_dimensions[row_idx].height = 60
        
        row_idx += 1

    wb.save(excel_file)
    engine_desc = "对比表格" if engine is None else f"{engine}报告"
    print(f"数据已填充到: {excel_file}，共添加 {len(data_dict)} 个模型（{engine_desc}）")


if __name__ == "__main__":
    # 示例数据
    example_data = {
        "DeepSeek-R1-Distill-Llama-70B": {
            "SigInfer_opencompass": "{'gsm8k': 74.5, 'failed': [], 'ARC-c': 84.5}\n{'mmlu': 90.526, 'ceval': 22.5}",
            "SigInfer_SGLang": "mmlu评测结果:['Average accuracy: 0.790','']\ngsm8k评测结果:['Accuracy: 0.935', 'Invalid: 0.000', 'Latency: 48.705 s', '']",
            "VLLM_opencompass": "{'gsm8k': 75.5, 'failed': [], 'ARC-c': 79.0}\n{'mmlu': 90.526, 'failed': [], 'ceval': 21.731}",
            "VLLM_SGLang": "mmlu评测结果:['Average accuracy: 0.790','']\ngsm8k评测结果:['Accuracy: 0.935', 'Invalid: 0.000', 'Latency: 41.991 s', '']"
        },
        "DeepSeek-R1-Distill-Qwen-32B": {
            "SigInfer_opencompass": "{'gsm8k': 72.0, 'failed': [], 'ARC-c': 72.0}\n{'mmlu': 86.842, 'ceval': 24.808}",
            "SigInfer_SGLang": "mmlu评测结果:['Average accuracy: 0.810','']\ngsm8k评测结果:['Accuracy: 0.870', 'Invalid: 0.000', 'Latency: 33.944 s', '']",
            "VLLM_opencompass": "{'gsm8k': 74.5, 'failed': [], 'ARC-c': 71.0}\n{'mmlu': 88.772, 'failed': [], 'ceval': 25.962}",
            "VLLM_SGLang": "mmlu评测结果:['Average accuracy: 0.809','']\ngsm8k评测结果:['Accuracy: 0.865', 'Invalid: 0.000', 'Latency: 31.630s', '']"
        },
        "Qwen2.5-32B-Instruct": {
            "SigInfer_opencompass": "{'gsm8k': 85.0, 'failed': [], 'ARC-c': 83.0}\n{'mmlu': 84.211, 'ceval': 61.923}",
            "SigInfer_SGLang": "mmlu评测结果:['Average accuracy: 0.832','']\ngsm8k评测结果:['Accuracy: 0.475', 'Invalid: 0.000', 'Latency: 56.580 s', '']",
            "VLLM_opencompass": "{'gsm8k': 84.5, 'failed': [], 'ARC-c': 83.0}\n{'mmlu': 85.088, 'failed': [], 'ceval': 59.231}",
            "VLLM_SGLang": "mmlu评测结果:['Average accuracy: 0.831','']\ngsm8k评测结果:['Accuracy: 0.715', 'Invalid: 0.000', 'Latency: 47.133 s', '']"
        },
        "Qwen2.5-72B-Instruct": {
            "SigInfer_opencompass": "{'gsm8k': 90.5, 'failed': [], 'ARC-c': 94.5}\n{'mmlu': 85.965, 'failed': [], 'ceval': 82.692}",
            "SigInfer_SGLang": "mmlu评测结果:['Average accuracy: 0.847','']\ngsm8k评测结果:['Accuracy: 0.245', 'Invalid: 0.000', 'Latency: 46.397 s', '']",
            "VLLM_opencompass": "{'gsm8k': 91.5, 'failed': [], 'ARC-c': 95.0}\n{'mmlu': 85.789, 'failed': [], 'ceval': 80.577}",
            "VLLM_SGLang": "mmlu评测结果:['Average accuracy: 0.845','']\ngsm8k评测结果:['Accuracy: 0.940', 'Invalid: 0.000', 'Latency: 34.137 s', '']"
        },
        "Meta-Llama-3.1-70B-Instruct": {
            "SigInfer_opencompass": "{'gsm8k': 89.0, 'failed': [], 'ARC-c': 94.0}\n{'mmlu': 81.579, 'ceval': 60.577}",
            "SigInfer_SGLang": "mmlu评测结果:['Average accuracy: 0.825','']\ngsm8k评测结果:['Accuracy: 0.935', 'Invalid: 0.000', 'Latency: 28.243 s', '']",
            "VLLM_opencompass": "{'gsm8k': 90.5, 'failed': [], 'ARC-c': 94.0}\n{'mmlu': 82.632, 'failed': [], 'ceval': 61.923}",
            "VLLM_SGLang": "mmlu评测结果:['Average accuracy: 0.825','']\ngsm8k评测结果:['Accuracy: 0.945', 'Invalid: 0.005', 'Latency: 24.806 s', '']"
        }
    }
    
    # 使用示例：
    print("=" * 80)
    print("示例 1: 生成对比报告（SigInfer vs VLLM）")
    print("=" * 80)
    comparison_file = "model_comparison_table.xlsx"
    generate_model_comparison_excel(comparison_file, engine=None)
    fill_model_comparison_data(comparison_file, example_data, engine=None)
    
    print("\n" + "=" * 80)
    print("示例 2: 生成SigInfer单独报告")
    print("=" * 80)
    siginfer_file = "siginfer_report.xlsx"
    generate_model_comparison_excel(siginfer_file, engine="SigInfer")
    fill_model_comparison_data(siginfer_file, example_data, engine="SigInfer")
    
    print("\n" + "=" * 80)
    print("示例 3: 生成VLLM单独报告")
    print("=" * 80)
    vllm_file = "vllm_report.xlsx"
    generate_model_comparison_excel(vllm_file, engine="VLLM_0.9.1")
    fill_model_comparison_data(vllm_file, example_data, engine="VLLM_0.9.1")
    
    print("\n" + "=" * 80)
    print("所有报告生成完成！")
    print("=" * 80)


from openpyxl import load_workbook

version = 
Model = 

# 加载 Excel 文件
file_path = f'./{version}/Supported_Model_list.xlsx'  # 替换为你的 Excel 文件路径
workbook = load_workbook(file_path)

# 选择工作表
sheet = workbook['Sheet1']  # 替换为实际工作表名

# 获取行数
row_count = sheet.max_row
print(f"总行数: {row_count}")

# 遍历行
for row in sheet.iter_rows(min_row=2, max_row=row_count, values_only=True):
    print(row)  # 每行数据以元组形式返回
    name = row[0]
    GPU = row[1]
    args = row[2]
    if name == Model:
        return args

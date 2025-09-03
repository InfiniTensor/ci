import sys
import re
from openpyxl import load_workbook

def main():
    if len(sys.argv) != 4:
        print("Usage: python SendMsgToBot <docker_image_version> <model_name> <log_file_path>")
        sys.exit(1)
    
    docker_image_version = sys.argv[1]
    model_name = sys.argv[2]
    log_file_path = sys.argv[3]
    
    # 读取日志文件内容
    try:
        with open(log_file_path, 'r', encoding='utf-8') as file:
            text = file.read()
    except FileNotFoundError:
        print(f"Error: Log file '{log_file_path}' not found. Please provide the correct file path.")
        exit(1)
    except Exception as e:
        print(f"Error reading log file: {e}")
        exit(1)

    # 定义正则表达式模式
    '''
    patterns = [
        # 匹配 Serving Benchmark Result 中的指标
        r"(?P<key>[A-Za-z\s\(\)/.-]+):\s+(?P<value>[0-9.]+)",
        # 匹配 Random Testing input=128, output=128
        r"(?P<key>input|output)=(?P<value>\d+)",
        # 匹配 Testing concurrency=1, prompts=4
        r"(?P<key>concurrency|prompts)=(?P<value>\d+)"
    ]
    '''
    
    # 定义正则表达式模式
    patterns = [
        # 匹配 Random Testing input 和 output
        r"Random Testing input=(\d+), output=(\d+)",
        # 匹配 Testing concurrency 和 prompts
        r"Testing concurrency=(\d+), prompts=(\d+)",
        # 匹配 Serving Benchmark Result 中的指标
        r"(?P<key>[A-Za-z\s\(\)/.-]+):\s+(?P<value>[0-9.]+)"
    ]

    # 上下文长度
    in_out_row_map={
        "128+128": 6,
        "128+1024": 15,
        "128+2048": 24,
        "1024+1024": 33,
        "2048+2048": 42,
        "4096+1024": 51,
        "1024+4096": 60,
        "30000+2048": 69,
        "126000+2048": 78
    }
    
    # 统计指标列名
    columns = [
        "Output Token Throughput",
        "Total Token Throughput",
        "Mean TTFT",
        "Median TTFT",
        "P99 TTFT",
        "Mean TPOT",
        "Median TPOT",
        "P99 TPOT",
        "Mean ITL",
        "Median ITL",
        "P99 ITL"
    ]

    # 创建列名到列索引的映射
    column_name_map = {column: index + 4 for index, column in enumerate(columns)}

    # 打印映射
    # for column, index in column_name_map.items():
    #    print(f"{column} => {index}")

    # 存储提取的结果
    results = {}
    current_config = {}

    # 对文本进行匹配
    matches = re.finditer(r"(Random Testing [^\n]+|Testing concurrency=[^\n]+|[\=]+ Serving Benchmark Result [\=]+.*?[\=]+)", text, re.DOTALL)
    for match in matches:
        section = match.group(0)
        
        # 匹配 input 和 output
        input_output_match = re.search(r"Random Testing input=(\d+), output=(\d+)", section)
        if input_output_match:
            current_config["input"] = int(input_output_match.group(1))
            current_config["output"] = int(input_output_match.group(2))
            in_out_length_key = f"{current_config['input']}+{current_config['output']}"
            results[in_out_length_key] = {}
            # print(section)
        
        # 匹配 concurrency 和 prompts
        concurrency_prompt_match = re.search(r"Testing concurrency=(\d+), prompts=(\d+)", section)
        if concurrency_prompt_match:
            current_config["concurrency"] = int(concurrency_prompt_match.group(1))
            current_config["prompts"] = int(concurrency_prompt_match.group(2))
            concurrency_key = f"{current_config['concurrency']}_{current_config['prompts']}"
            results[in_out_length_key][concurrency_key] = {}
            # print(section)
        
        # 匹配基准测试结果
        benchmark_match = re.search(r"[\=]+ Serving Benchmark Result [\=]+.*?[\=]+", section, re.DOTALL)
        if benchmark_match:
            metrics = re.finditer(r"(?P<key>[A-Za-z0-9\s\(\)/]+):\s+(?P<value>[0-9.]+)", benchmark_match.group(0))
            for metric in metrics:
                key = metric.group("key").strip()
                value = float(metric.group("value"))
                # print(f"{key}=>{value}")
                results[in_out_length_key][concurrency_key][key] = value

    # 打印结果
    for in_out_length_key, concurrency_map in results.items():
        print(f"{in_out_length_key}:")
        for concurrency_key, metrics in concurrency_map.items():
            print(f"  {concurrency_key}:")
            for key, value in metrics.items():
                print(f"    {key}: {value}")
    
    # 以字典形式返回
    # print("\nResults as dictionary:")
    # print(results)
    
    # 加载已有Excel模板
    wb = load_workbook('report_template.xlsx')
    ws = wb.active  # 或者用 wb["Sheet1"]
    
    for in_out_length_key, concurrency_map in results.items():
        start_row = in_out_row_map[in_out_length_key]
        for conncurrency_key, metrics in concurrency_map.items():
            for key, value in metrics.items():
                ws.cell(row=start_row, column=column_name_map["Output Token Throughput"]).value = metrics["Output token throughput (tok/s)"]
                ws.cell(row=start_row, column=column_name_map["Total Token Throughput"]).value = metrics["Total Token throughput (tok/s)"]
                ws.cell(row=start_row, column=column_name_map["Mean TTFT"]).value = metrics["Mean TTFT (ms)"]
                ws.cell(row=start_row, column=column_name_map["Median TTFT"]).value = metrics["Median TTFT (ms)"]
                ws.cell(row=start_row, column=column_name_map["P99 TTFT"]).value = metrics["P99 TTFT (ms)"]
                ws.cell(row=start_row, column=column_name_map["Mean TPOT"]).value = metrics["Mean TPOT (ms)"]
                ws.cell(row=start_row, column=column_name_map["Median TPOT"]).value = metrics["Median TPOT (ms)"]
                ws.cell(row=start_row, column=column_name_map["P99 TPOT"]).value = metrics["P99 TPOT (ms)"]
                ws.cell(row=start_row, column=column_name_map["Mean ITL"]).value = metrics["Mean ITL (ms)"]
                ws.cell(row=start_row, column=column_name_map["Median ITL"]).value = metrics["Median ITL (ms)"]
                ws.cell(row=start_row, column=column_name_map["P99 ITL"]).value = metrics["P99 ITL (ms)"]
            start_row += 1

    # 保存新文件
    wb.save('./report/' + model_name + '.xlsx')
 
if __name__ == "__main__":
    main()

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def generate_excel(model_name, exec_cmd, test_cmd, context_lengths, batch_sizes, multiplier, output_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ascend"

    # 样式
    subject_fill = PatternFill("solid", fgColor="FF99CC")  # 表头粉色
    header_fill = PatternFill("solid", fgColor="FFD966")  # 表头黄色
    blue_fill = PatternFill("solid", fgColor="9BC2E6")    # 浅蓝
    light_blue_fill = PatternFill("solid", fgColor="D9E1F2")  # 更浅蓝
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 顶部 3 行：模型、启动命令、测试命令
    top_info = [
        ("模型", model_name),
        ("启动命令", exec_cmd),
        ("测试命令", test_cmd)
    ]

    col_count = 15  # 总列数（SigInfer + 14列指标）
    row_idx = 1
    for title, content in top_info:
        ws.cell(row=row_idx, column=1, value=title)
        ws.cell(row=row_idx, column=1).font = Font(bold=True)
        ws.cell(row=row_idx, column=1).alignment = left_align
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=1).fill = subject_fill

        # 合并后面的列写内容
        ws.merge_cells(start_row=row_idx, start_column=2,
                       end_row=row_idx, end_column=col_count)
        ws.cell(row=row_idx, column=2, value=content)
        ws.cell(row=row_idx, column=2).alignment = left_align
        # ws.cell(row=row_idx, column=2).border = thin_border
        row_idx += 1

    # 二级表头
    # 第 4 行：大类
    top_headers = [
        (4, 5, "Serving Benchmark Result"),
        (6, 8, "Time to First Token"),
        (9, 11, "Time per Output Token"),
        (12, 14, "Inter-token Latency")
    ]

    # 第 5 行：细项
    sub_headers = [
        "Inference Engine", "上下文长度", "Batch",
        "Output token throughput", "Total Token throughput",
        "Mean TTFT", "Median TTFT", "P99 TTFT",
        "Mean TPOT", "Median TPOT", "P99 TPOT",
        "Mean ITL", "Median ITL", "P99 ITL"
    ]

    # 写入第 4 行并合并
    for start_col, end_col, title in top_headers:
        ws.merge_cells(start_row=4, start_column=start_col,
                       end_row=4, end_column=end_col)
        cell = ws.cell(4, start_col, value=title)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border
        # 对于被合并区域内其余单元格也加边框
        for col in range(start_col + 1, end_col + 1):
            c = ws.cell(4, col)
            c.fill = header_fill
            c.border = thin_border

    # 将前 3 列在第 4~5 行纵向合并
    for col, label in zip([1, 2, 3], ["Inference Engine", "上下文长度", "Batch"]):
        ws.merge_cells(start_row=4, start_column=col,
                       end_row=5, end_column=col)
        c = ws.cell(4, col, value=label)
        c.alignment = center_align
        # 让文字视觉上靠近第 5 行
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="bottom")
        c.font = Font(bold=True)
        c.border = thin_border
        # 对于被合并区域内其余单元格也加边框
        start_row = 4
        end_row = 5
        for row in range(start_row + 1, end_row + 1):
            c = ws.cell(row, col)
            c.border = thin_border

    # 写入第 5 行细项（从第 4 列开始，因为 1~3 列已与上一行合并）
    for offset, title in enumerate(sub_headers[3:], start=4):
        cell = ws.cell(5, offset, value=title)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = center_align
        cell.border = thin_border

    # 冻结表头以上
    ws.freeze_panes = "A6"

    row_index = 6  # 数据从第 6 行开始
    for context in context_lengths:
        start_row = row_index  # 合并上下文开始行
        for i, batch in enumerate(batch_sizes):
            ws.cell(row=row_index, column=1, value="SigInfer")  # 第一列
            ws.cell(row=row_index, column=2, value=context)     # 第二列
            ws.cell(row=row_index, column=3,
                    value=f"{batch} (num_prompt={int(batch * multiplier)})")

            # 其他列预置为空并设为浮点显示两位小数
            for col in range(4, 15):
                c = ws.cell(row=row_index, column=col, value=None)
                c.data_type = "n"
                c.number_format = "0.00"
                c.alignment = right_align

            # 背景色交替
            row_fill = blue_fill if row_index % 2 == 0 else light_blue_fill

            # 应用样式
            for col in range(1, 15):
                cell = ws.cell(row=row_index, column=col)
                if col == 1 or col == 2:
                    cell.alignment = center_align
                elif col == 3:
                    cell.alignment = left_align
                    cell.fill = row_fill
                else:
                    cell.fill = row_fill
                cell.border = thin_border

            row_index += 1

        # 合并上下文长度单元格
        ws.merge_cells(start_row=start_row, start_column=2,
                       end_row=row_index - 1, end_column=2)
        ws.cell(start_row, 2).alignment = center_align
        ws.cell(start_row, 2).font = Font(bold=True)

    # 合并 SigInfer 整列
    ws.merge_cells(start_row=6, start_column=1,
                   end_row=row_index - 1, end_column=1)
    ws.cell(6, 1).alignment = center_align
    ws.cell(6, 1).font = Font(bold=True)

    # 自动调整列宽
    for col in range(1, 15):
        if col == 3 or col == 4 or col == 5:
            ws.column_dimensions[get_column_letter(col)].width = 26
        else:
            ws.column_dimensions[get_column_letter(col)].width = 19

    wb.save(output_file)
    print(f"Excel 文件已生成: {output_file}")


def fill_benchmark_results(excel_file, benchmark_data, context_lengths, batch_sizes):
    """
    填充基准测试结果到Excel表格

    Args:
        excel_file: Excel文件路径
        benchmark_data: 字典格式的基准测试数据
                      格式: {(context_length, batch_size): {
                           'output_token_throughput': float,
                           'total_token_throughput': float,
                           'mean_ttft': float,
                           'median_ttft': float,
                           'p99_ttft': float,
                           'mean_tpot': float,
                           'median_tpot': float,
                           'p99_tpot': float,
                           'mean_itl': float,
                           'median_itl': float,
                           'p99_itl': float
                       }}
        context_lengths: 上下文长度列表
        batch_sizes: 批次大小列表
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 数据列映射 (列索引 -> 数据键)
    data_columns = {
        4: 'Output token throughput (tok/s)',
        5: 'Total Token throughput (tok/s)',
        6: 'Mean TTFT (ms)',
        7: 'Median TTFT (ms)',
        8: 'P99 TTFT (ms)',
        9: 'Mean TPOT (ms)',
        10: 'Median TPOT (ms)',
        11: 'P99 TPOT (ms)',
        12: 'Mean ITL (ms)',
        13: 'Median ITL (ms)',
        14: 'P99 ITL (ms)'
    }

    row_index = 6  # 数据从第6行开始

    for context in context_lengths:
        for batch in batch_sizes:
            # 查找对应的基准测试数据
            key = (context, batch)
            if key in benchmark_data:
                data = benchmark_data[key]
                # 填充数据列 (第4-14列)
                for col, data_key in data_columns.items():
                    if data_key in data:
                        cell = ws.cell(row=row_index, column=col,
                                       value=data[data_key])
                        cell.number_format = "0.00"
            row_index += 1

    wb.save(excel_file)
    print(f"基准测试结果已填充到: {excel_file}")


def fill_benchmark_results_from_list(excel_file, results_list, context_lengths, batch_sizes):
    """
    从结果列表填充基准测试数据

    Args:
        excel_file: Excel文件路径
        results_list: 结果列表，按顺序对应每个(context_length, batch_size)组合
                     每个元素是包含11个数值的列表或元组:
                     [output_token_throughput, total_token_throughput, 
                      mean_ttft, median_ttft, p99_ttft,
                      mean_tpot, median_tpot, p99_tpot,
                      mean_itl, median_itl, p99_itl]
        context_lengths: 上下文长度列表
        batch_sizes: 批次大小列表
    """
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    # 数据列索引
    data_columns = [4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14]

    row_index = 6  # 数据从第6行开始
    result_idx = 0

    for context in context_lengths:
        for batch in batch_sizes:
            if result_idx < len(results_list):
                result = results_list[result_idx]

                # 填充数据列
                for i, col in enumerate(data_columns):
                    if i < len(result) and result[i] is not None:
                        cell = ws.cell(row=row_index, column=col,
                                       value=float(result[i]))
                        cell.number_format = "0.00"

            row_index += 1
            result_idx += 1

    wb.save(excel_file)
    print(f"基准测试结果已填充到: {excel_file}")


if __name__ == "__main__":
    context_lengths = ["128+128", "128+1024", "128+2048",
                       "1024+1024", "2048+2048", "4096+1024", "1024+4096"]
    batch_sizes = [1, 5, 10, 20, 50, 100, 150]

    # 生成Excel模板
    generate_excel(context_lengths, batch_sizes, 4, "benchmark_template.xlsx")

    # 示例：使用字典格式填充数据
    example_benchmark_data = {
        ("128+128", 1): {
            'output_token_throughput': 76.26,
            'total_token_throughput': 151.92,
            'mean_ttft': 40.15,
            'median_ttft': 38.20,
            'p99_ttft': 65.30,
            'mean_tpot': 12.50,
            'median_tpot': 12.10,
            'p99_tpot': 18.90,
            'mean_itl': 8.20,
            'median_itl': 7.80,
            'p99_itl': 15.40
        },
        ("128+128", 5): {
            'output_token_throughput': 85.30,
            'total_token_throughput': 170.60,
            'mean_ttft': 45.20,
            'median_ttft': 43.10,
            'p99_ttft': 72.80,
            'mean_tpot': 11.80,
            'median_tpot': 11.40,
            'p99_tpot': 17.20,
            'mean_itl': 7.50,
            'median_itl': 7.20,
            'p99_itl': 14.10
        }
    }

    # 填充示例数据
    # fill_benchmark_results("benchmark_template.xlsx", example_benchmark_data, context_lengths, batch_sizes)

    # 示例：使用列表格式填充数据
    example_results_list = [
        [76.26, 151.92, 40.15, 38.20, 65.30, 12.50,
            12.10, 18.90, 8.20, 7.80, 15.40],
        [85.30, 170.60, 45.20, 43.10, 72.80, 11.80,
            11.40, 17.20, 7.50, 7.20, 14.10],
        # 可以继续添加更多结果...
    ]

    # 填充示例数据
    # fill_benchmark_results_from_list("benchmark_template.xlsx", example_results_list, context_lengths, batch_sizes)

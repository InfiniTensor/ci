from openpyxl import load_workbook
import re
import os
import sys


def main():
    if len(sys.argv) != 3:
        print("Usage: python script_generator_for_MindIE.py <test_type> <version>")
        sys.exit(1)
    
    test_type = sys.argv[1]
    version = sys.argv[2]
    
    curr_dir = os.getcwd()
    
    # 加载 Excel 文件
    file_path = f'{curr_dir}/{version}/model_list.xlsx'  # 替换为你的 Excel 文件路径
    workbook = load_workbook(file_path)

    # 选择工作表
    sheet = workbook['MindIE']

    # 获取行数
    row_count = sheet.max_row
    print(f"总行数: {row_count}")
    
    target_file = ""
    src_code = ""
    
    if test_type == "Smoke":
        port_num = "$((6543+${JOB_COUNT}))"
        target_file = "job_executor_for_SmokeTest.sh"
    elif test_type == "Performance":
        port_num = "$((8765+${JOB_COUNT}))"
        target_file = "job_executor_for_PerformanceTest.sh"
    elif test_type == "Stability":
        port_num = "$((8000+${JOB_COUNT}))"
        target_file = "job_executor_for_StabilityTest.sh"
    elif test_type == "Accuracy":
        port_num = "$((9701+${JOB_COUNT}))"
        target_file = "job_executor_for_AccuracyTest.sh"
    
    start = True
    for row in sheet.iter_rows(min_row=2, max_row=row_count, values_only=True):
        # print(row)  # 每行数据以元组形式返回
        name = row[0]
        GPU = row[1]
        args = row[2]
        args = args.split('\n')[0]
        
        pattern = "--tokenizer\s+(\S+)"
        match = re.search(pattern, args)
        if match:
            model_weight_path = match.group(1)
        else:
            model_weight_path = ""
        
        if start:
            src_code += f"    if [ $MODEL == \"{name}\" ]; then\n"
            start = False
        else:
            src_code += f"    elif [ $MODEL == \"{name}\" ]; then\n"
        src_code += f"        MODEL_WEIGHT_PATH=\"{model_weight_path}\"\n"
    src_code += "    fi\n"
    
    # print(src_code)

    template_file = "job_executor_template_for_MindIE.sh"

    try:
        with open(f"{curr_dir}/{template_file}", 'r', encoding='utf-8') as file:
            lines = file.readlines()
        
        line_num = 0
        for line in lines:
            if "<<<generated source code>>>" in line:
                lines[line_num] = src_code
            elif "<<<TEST_TYPE>>>" in line:
                if test_type == "Smoke":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "SmokeTest")
                elif test_type == "Performance":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "PerformanceTest")
                elif test_type == "Accuracy":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "AccuracyTest")
                elif test_type == "Stability":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "StabilityTest")
            elif "<<<PORT>>>" in line:
                lines[line_num] = line.replace("<<<PORT>>>", port_num)
            line_num += 1

        with open(f"{curr_dir}/{target_file}", 'w') as file:
            file.write(''.join(lines))
    except FileNotFoundError:
        print(f"Error: Log file '{curr_dir}/{template_file}' not found.")
    except Exception as e:
        print(f"Error reading file: {str(e)}")
    
    os.system(f"chmod 777 {target_file}")

if __name__ == "__main__":
    main()

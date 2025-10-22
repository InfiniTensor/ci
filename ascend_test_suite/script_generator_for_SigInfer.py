from openpyxl import load_workbook
import re
import os
import sys


def main():
    verison = ""
    if len(sys.argv) != 3:
        print("Usage: python script_generator <test_type> <version>")
        sys.exit(1)
    
    test_type = sys.argv[1]
    version = sys.argv[2]
    
    curr_dir = "/home/s_limingge/ascend_test_suite"

    # 加载 Excel 文件
    file_path = f'{curr_dir}/{version}/model_list.xlsx'  # 替换为你的 Excel 文件路径
    workbook = load_workbook(file_path)

    # 选择工作表
    sheet = workbook['Ascend']

    # 获取行数
    row_count = sheet.max_row
    print(f"总行数: {row_count}")
    
    target_file = ""
    src_code = ""
    
    if test_type == "Smoke":
        src_code += "PORT=$((6543+${JOB_COUNT}))\n"
        src_code += "PROMETHEUS_PORT=$((8321+${JOB_COUNT}))\n"
        src_code += "MASTER_PORT=$((8438+${JOB_COUNT}))\n"
        src_code += "LOG_NAME=\"server_log_SmokeTest_$(date +'%Y%m%d_%H%M%S').log\"\n\n"
        target_file = "job_executor_for_SmokeTest.sh"
    elif test_type == "Performance":
        src_code += "PORT=$((8765+${JOB_COUNT}))\n"
        src_code += "PROMETHEUS_PORT=$((28765+${JOB_COUNT}))\n"
        src_code += "MASTER_PORT=$((9642+${JOB_COUNT}))\n"
        src_code += "LOG_NAME=\"server_log_PerformanceTest_$(date +'%Y%m%d_%H%M%S').log\"\n\n"
        target_file = "job_executor_for_PerformanceTest.sh"
    elif test_type == "Stability":
        src_code += "PORT=$((8000+${JOB_COUNT}))\n"
        src_code += "PROMETHEUS_PORT=$((28880+${JOB_COUNT}))\n"
        src_code += "MASTER_PORT=$((9032+${JOB_COUNT}))\n"
        src_code += "LOG_NAME=\"server_log_StabilityTest_$(date +'%Y%m%d_%H%M%S').log\"\n\n"
        target_file = "job_executor_for_StabilityTest.sh"
    elif test_type == "Accuracy":
        src_code += "PORT=$((9701+${JOB_COUNT}))\n"
        src_code += "PROMETHEUS_PORT=$((25771+${JOB_COUNT}))\n"
        src_code += "MASTER_PORT=$((22642+${JOB_COUNT}))\n"
        src_code += "LOG_NAME=\"server_log_AccuracyTest_$(date +'%Y%m%d_%H%M%S').log\"\n\n"
        target_file = "job_executor_for_AccuracyTest.sh"
    
    start = True
    for row in sheet.iter_rows(min_row=2, max_row=row_count, values_only=True):
        # print(row)  # 每行数据以元组形式返回
        name = row[0]
        GPU = row[1]
        args = row[2]
        args = args.split('\n')[0]
        
        result = re.sub(r"^.*docker\.xcoresigma\.com/docker/siginfer-aarch64-ascend\:\S+", "", args)
        result = re.sub(r"--swap-space\s+\d+", "$SWAP_SPACE_OPTION", result)
        result = re.sub(r"--prometheus-port\s+\d+", "--prometheus-port $PROMETHEUS_PORT", result)
        result = re.sub(r"--port\s+\d+", "--port $PORT", result)
        result = re.sub(r"--master-addr\s+\S+", "--master-addr $MASTER_IP", result)
        result = re.sub(r"--node-rank\s+\d+", "--node-rank $NODE_RANK", result)
        result = re.sub(r"--schedule-policy\s+\S+", "--schedule-policy $SCHEDULE_POLICY", result)
        result = re.sub(r"--master-port\s+\d+", "--master-port $MASTER_PORT", result)
        result = re.sub(r"--use-prefix-cache", "$USE_PREFIX_CACHE", result)
        result += " --gpu-memory-utilization 0.95"
        result += " --prometheus-port $PROMETHEUS_PORT"
        
        if start:
            src_code += f"if [ $MODEL == \"{name}\" ]; then\n"
            start = False
        else:
            src_code += f"elif [ $MODEL == \"{name}\" ]; then\n"
        src_code += "    echo \"SIG_LOG_LEVEL='warn,console_logger=info' ./start.sh"
        src_code += result
        src_code += "\"\n"
        
        src_code += "    EXEC_COMMAND+=\""
        src_code += result
        src_code += " > $LOG_NAME 2>&1 &\"\n"
        
    src_code += "fi\n"

    # print(src_code)

    template_file = "job_executor_template_for_SigInfer.sh"

    try:
        with open(f"{curr_dir}/{template_file}", 'r', encoding='utf-8') as file:
            lines = file.readlines()
        
        line_num = 0
        for line in lines:
            if "<<<generated source code>>>" in line:
                lines[line_num] = src_code
                with open(f"{curr_dir}/{target_file}", 'w') as file:
                    file.write(''.join(lines))
                # print(lines)
                break
            elif "<<<TEST_TYPE>>>" in line:
                if test_type == "Smoke":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "SmokeTest")
                elif test_type == "Performance":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "PerformanceTest")
                elif test_type == "Accuracy":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "AccuracyTest")
                elif test_type == "Stability":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "StabilityTest")
            line_num += 1
    except FileNotFoundError:
        print(f"Error: Log file '{curr_dir}/{template_file}' not found.")
    except Exception as e:
        print(f"Error reading file: {str(e)}")
    
    os.system(f"chmod 777 {target_file}")

if __name__ == "__main__":
    main()

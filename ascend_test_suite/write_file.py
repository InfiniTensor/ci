import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment

parser = argparse.ArgumentParser(description="写入结果文件")
parser.add_argument('--file', type=str)
parser.add_argument('--framework', type=str)

args=parser.parse_args()
# 读取 txt 文件，按 '+' 分割列
input_file = args.file
output_file = args.framework

output_file = f"{output_file}_result.xlsx"

wb = Workbook()
ws = wb.active

with open(input_file, "r", encoding="utf-8") as file:
    for row_idx, line in enumerate(file, start=1):  # 从第1行开始
        columns = line.strip().split('+')  # 去除换行符并按 '+' 分割
        print(columns)
        if len(columns[1].split('} {')) == 2:
            columns[1] = columns[1].replace("} {", "} \n{")
        # 在"gsm8k"前添加换行符
        if 'gsm8k' in columns[2]:
            columns[2] = columns[2].replace("gsm8k", "\ngsm8k")
        for col_idx, value in enumerate(columns, start=1):  # 从第1列开始
            ws.cell(row=row_idx, column=col_idx, value=value)
            
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, str) and '\n' in cell.value:
            cell.alignment = Alignment(wrap_text=True)
            
ws.column_dimensions['B'].width=50
ws.column_dimensions['C'].width=78

wb.save(output_file)
print(f"数据已写入 {output_file}")

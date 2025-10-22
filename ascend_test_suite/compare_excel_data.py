#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel性能数据比较脚本
读取Excel文件中的性能数据并比较差值是否在20%以内
"""

from openpyxl import load_workbook
import sys
from pathlib import Path
from SendMsgToBot import send_summary_to_server

def read_excel_data(file_path):
    """读取Excel文件的所有sheet"""
    print(f"\n正在读取文件: {file_path}")
    
    # 读取Excel文件
    wb = load_workbook(file_path, data_only=True)
    
    print(f"发现 {len(wb.sheetnames)} 个sheet:")
    for i, sheet_name in enumerate(wb.sheetnames, 1):
        print(f"  {i}. {sheet_name}")
    
    # 读取所有sheet的数据
    sheets_data = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 读取数据到列表
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)
        
        sheets_data[sheet_name] = {
            'worksheet': ws,
            'data': data
        }
        
        print(f"\nSheet '{sheet_name}':")
        print(f"  - 行数: {ws.max_row}, 列数: {ws.max_column}")
        if len(data) > 0:
            print(f"  - 列头: {data[0]}")
            print(f"  - 前3行数据:")
            for i, row in enumerate(data[:3], 1):
                print(f"    行{i}: {row}")
    
    return sheets_data

def is_numeric(value):
    """检查值是否为数字"""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return not (value != value)  # 检查是否为NaN
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False

def compare_two_sheets(sheet1_data, sheet2_data, sheet1_name, sheet2_name, threshold, docker_image_version, model_name):
    """比较两个sheet的数据，检查数值的差值是否在阈值内"""
    print(f"\n{'='*100}")
    print(f"比较 '{sheet1_name}' 和 '{sheet2_name}' 的性能测试数据")
    print(f"{'='*100}")
    
    data1 = sheet1_data['data']
    data2 = sheet2_data['data']
    
    if len(data1) == 0 or len(data2) == 0:
        print("警告: 其中一个sheet为空")
        return
    
    # 找到性能指标表头行（包含"Output token throughput"等）
    header_row_idx = -1
    metric_headers = []
    
    for idx, row in enumerate(data1):
        if any(cell and 'Output token throughput' in str(cell) for cell in row):
            metric_headers = list(row)  # 转换为列表以便修改
            header_row_idx = idx
            print(f"\n找到性能指标表头行: 第{idx+1}行")
            print(f"性能指标列: {[str(h) for h in metric_headers if h]}")
            break
    
    if header_row_idx == -1:
        print("警告: 未找到性能指标表头，将从第6行开始比较")
        header_row_idx = 5  # 默认从第6行开始
        metric_headers = list(data1[header_row_idx]) if len(data1) > header_row_idx else []
    
    # 找到最长的列名并填充其他列名以实现对齐
    if metric_headers:
        # 计算最长的列名长度
        max_length = max(len(str(h)) for h in metric_headers if h)
        # 给其他列名尾部填充空格
        metric_headers = [str(h).ljust(max_length) if h else h for h in metric_headers]
        print(f"列名对齐完成，最大长度: {max_length}")
        print(f"对齐后的列名: {[repr(h) for h in metric_headers if h]}")
    
    # 比较结果
    comparison_results = []
    exceed_threshold_count = 0
    total_comparisons = 0
    
    # 从表头的下一行开始比较（跳过元数据和表头）
    start_row = header_row_idx + 1
    min_rows = min(len(data1), len(data2))
    min_cols = min(len(data1[0]) if data1 else 0, len(data2[0]) if data2 else 0)
    
    print(f"\n开始比较性能数据: 从第{start_row+1}行到第{min_rows}行")
    
    for row_idx in range(start_row, min_rows):
        # 获取上下文长度和Batch信息（用于标识这一行的测试配置）
        context_len = data1[row_idx][1] if len(data1[row_idx]) > 1 and data1[row_idx][1] != None else context_len
        batch = data1[row_idx][2] if len(data1[row_idx]) > 2 else None
        
        # 去掉batch值中的括号部分，如 "100 (num_prompt=400) " -> "100"
        if batch and '(' in str(batch):
            batch = str(batch).split('(')[0].strip()
        
        # 只比较有效的测试数据行（有上下文长度或batch信息）
        if not context_len and not batch:
            continue
        
        row_label = f"{context_len or ''} | {batch or ''}".strip(' |')
        
        # 从第4列开始比较性能数值（前3列通常是：Inference Engine, 上下文长度, Batch）
        for col_idx in range(3, min_cols):
            val1 = data1[row_idx][col_idx] if col_idx < len(data1[row_idx]) else None
            val2 = data2[row_idx][col_idx] if col_idx < len(data2[row_idx]) else None
            
            # 只比较数值
            if is_numeric(val1) and is_numeric(val2):
                val1_float = float(val1)
                val2_float = float(val2)
                
                if val1_float == 0:
                    continue
                
                # 计算百分比差异
                diff = abs(val2_float - val1_float)
                percent_diff = (diff / abs(val1_float)) * 100
                
                within_threshold = percent_diff <= (threshold * 100)
                total_comparisons += 1
                
                if not within_threshold:
                    exceed_threshold_count += 1
                
                # 获取指标名称
                metric_name = metric_headers[col_idx] if col_idx < len(metric_headers) and metric_headers[col_idx] else f"列{col_idx}"
                
                result = {
                    '测试配置': row_label,
                    '性能指标': str(metric_name),
                    f'{sheet1_name}_值': round(val1_float, 4),
                    f'{sheet2_name}_值': round(val2_float, 4),
                    '差值': round(diff, 4),
                    '百分比差异(%)': round(percent_diff, 2),
                    f'在{int(threshold*100)}%以内': '✓' if within_threshold else '✗'
                }
                comparison_results.append(result)
    
    # 打印结果
    if comparison_results:
        print(f"\n性能数据比较结果 (共 {len(comparison_results)} 条):")
        print(f"\n{'='*170}")
        
        # 打印表头 - 更清晰的列名，使用更合理的列宽
        print(f"{'测试配置':<25} {'性能指标':<30} {sheet1_name+'_值':<15} {sheet2_name+'_值':<15} {'差值':<12} {'百分比差异(%)':<15} {'在阈值内':<8}")
        print(f"{'-'*140}")
        
        # 打印所有数据
        for result in comparison_results:
            print(f"{result['测试配置']:<25} {result['性能指标']:<30} {result[f'{sheet1_name}_值']:<15} {result[f'{sheet2_name}_值']:<15} {result['差值']:<12} {result['百分比差异(%)']:<15} {result[f'在{int(threshold*100)}%以内']:<8}")
        
        summary = ""
        # 统计信息
        print(f"\n{'='*170}")
        print(f"\n📊 统计信息:")
        summary += f"\n📊 统计信息:\n"
        print(f"  ✓ 总比较次数: {total_comparisons}")
        summary += f"  ✓ 总比较次数: {total_comparisons}\n"
        print(f"  ✓ 在{int(threshold*100)}%阈值内: {total_comparisons - exceed_threshold_count} ({(total_comparisons - exceed_threshold_count)/total_comparisons*100:.1f}%)")
        summary += f"  ✓ 在{int(threshold*100)}%阈值内: {total_comparisons - exceed_threshold_count} ({(total_comparisons - exceed_threshold_count)/total_comparisons*100:.1f}%)\n"
        print(f"  ✗ 超出阈值: {exceed_threshold_count} ({exceed_threshold_count/total_comparisons*100:.1f}%)")
        summary += f"  ✗ 超出阈值: {exceed_threshold_count} ({exceed_threshold_count/total_comparisons*100:.1f}%)\n"
        
        if exceed_threshold_count > 0:
            print(f"\n⚠️  警告: 有 {exceed_threshold_count} 条性能数据的差值超过 {int(threshold*100)}%!")
            summary += f"\n⚠️  警告: 有 {exceed_threshold_count} 条性能数据的差值超过 {int(threshold*100)}%!\n"
            print(f"\n超出阈值的数据详情:")
            summary += f"\n超出阈值的数据详情:\n"
            print(f"{'-'*140}")
            summary += f"{'-'*145}\n"
            print(f"{'测试配置':<25} {'性能指标':<30} {sheet1_name+'_值':<15} {sheet2_name+'_值':<15} {'差值':<12} {'百分比差异(%)':<15}")
            summary += f"{'测试配置':<25} {'性能指标':<30} {sheet1_name+'_值':<15} {sheet2_name+'_值':<15} {'差值':<12} {'百分比差异(%)':<15}\n"
            print(f"{'-'*140}")
            summary += f"{'-'*145}\n"
            for result in comparison_results:
                if result[f'在{int(threshold*100)}%以内'] == '✗':
                    print(f"{result['测试配置']:<25} {result['性能指标']:<30} {result[f'{sheet1_name}_值']:<15} {result[f'{sheet2_name}_值']:<15} {result['差值']:<12} {result['百分比差异(%)']:<15}")
                    summary += f"{result['测试配置']:<25} {result['性能指标']:<30} {result[f'{sheet1_name}_值']:<15} {result[f'{sheet2_name}_值']:<15} {result['差值']:<12} {result['百分比差异(%)']:<15}\n"
            summary += f"{'-'*145}\n"
        else:
            print(f"\n✅ 太棒了！所有性能数据的差值都在 {int(threshold*100)}% 以内!")
            summary += f"\n✅ 太棒了！所有性能数据的差值都在 {int(threshold*100)}% 以内!"
        
        print("\n")
        # 向飞书机器人发送测试结果报告
        send_summary_to_server("siginfer-aarch64-ascend:" + docker_image_version, model_name, summary)
        
        return comparison_results
    else:
        print("⚠️  没有找到可比较的性能数值数据")
        return None

def main():
    import sys
    
    # 文件路径
    if len(sys.argv) < 5:
        print("Usage: python compare_excel_data.py <docker_image_version> <model_name> <file_path1> <file_path2>")
        sys.exit(1)
    
    docker_image_version = sys.argv[1]
    model_name = sys.argv[2]
    file_path1 = sys.argv[3]
    file_path2 = sys.argv[4]
    
    # 读取Excel数据文件
    print(f"\n文件1: {file_path1}")
    sheets_data1 = read_excel_data(file_path1)
    
    print(f"\n文件2: {file_path2}")
    sheets_data2 = read_excel_data(file_path2)
    
    # 获取sheet名称
    sheet_names1 = list(sheets_data1.keys())
    sheet_names2 = list(sheets_data2.keys())
    
    # 比较两个sheet
    compare_two_sheets(
        sheets_data1[sheet_names1[0]], 
        sheets_data2[sheet_names2[0]], 
        f"文件1-{sheet_names1[0]}" if file_path1 != file_path2 else "基准",
        f"文件2-{sheet_names2[0]}" if file_path1 != file_path2 else "对比",
        0.20,
        docker_image_version,
        model_name
    )
    
    return sheets_data1, sheets_data2

if __name__ == "__main__":
    sheets_data = main()

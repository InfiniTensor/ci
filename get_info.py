import os
from logging import log
import json
from datetime import datetime
import pytz
import pandas as pd
import os
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment,Font
import argparse
from dynaconf import Dynaconf

parser = argparse.ArgumentParser(description="汇总结果写入文件")
parser.add_argument('--file', type=str)
parser.add_argument('--email', type=str)
parser.add_argument('--env', type=str)
parser.add_argument('--url', type=str)
parser.add_argument('--model', type=str)

args=parser.parse_args()
file_name = args.file
emails = args.email
env = args.env
url = args.url
model = args.model


if ',' in emails:
    emails = emails.split(',')
tz = pytz.timezone('Asia/Shanghai')
# time_str = "2024-05-27"
time_str = datetime.now(tz).strftime("%Y_%m_%d %H:%M:%S").split(' ')[0]
# 获取当前文件绝对路径
script_path = os.path.abspath(__file__)
# 获取ci_perf所需相对路径
pro_dir_path = os.path.dirname(script_path)
directory = f'{pro_dir_path}/allure-report/data/test-cases/'

def get_json_data(file_path):
    """
    读取json文件中的数据
    :param file_path: 文件路径
    :return:
    """
    if os.path.exists(file_path):
        files = os.path.splitext(file_path)
        filename, suffix = files  # 获取文件后缀
        if suffix == '.json':
            with open(file_path, 'r', encoding="utf-8") as fp:
                data = json.load(fp)
        else:
            log.error('文件后缀名错误')
    else:
        log.error('文件路径不存在')
    return data

cases_data = []

def get_file_contents(directory):
    
    for filename in os.listdir(directory):
        file_path = directory  + filename
        case_data = get_json_data(file_path )
        cases_data.append(case_data)
        # break

cases=[]
case_name_list=[]
case_status_list=[]
pass_num = 0
skip_num = 0
failed_num = 0 
def get_case_info():
    global pass_num
    global skip_num
    global failed_num
    get_file_contents(directory)
    for case in cases_data:

        case_name = case['name']
        case_status = case['status']
        case = {}
        case['name'] = case_name
        case['status'] = case_status
        if case_name in case_name_list:
            continue
        else:
            cases.append(case)
        if case_status == 'passed':
            pass_num += 1
        elif case_status == 'skipped':
            skip_num += 1
        else:
            failed_num +=1

get_case_info()
cases_new = sorted(cases,key=lambda x: x["name"])

for case in cases_new:
    case_name_list.append(case['name'])
    case_status_list.append(case['status'])
writeData = {  # 用字典设置DataFrame所需数据
    '用例名称': case_name_list,
    '运行结果': case_status_list,
}
fwrite = pd.DataFrame(writeData) 
# fwrite['首token时延加速比'] = fwrite.eval("`首token时延（s）2`/`首token时延（s）1`")
path=f'./{file_name}.xlsx'
fwrite.to_excel(path, index=False,engine='openpyxl')
wb=load_workbook(path)
for sheet in wb.sheetnames:#对Excel工作簿里面每个sheet表循环
    ws=wb[sheet]
    lks=[]#用来存储列宽数值
    for i in range(1,ws.max_column+1):
        lk=1#给一个初始列宽数值1
        for j in range(1,ws.max_row+1):#对每一列进行循环
            sz=ws.cell(row=j,column=i).value#获取每一列单元格的内容
            if isinstance(sz, (int, float)):#判断单元格是不是数字格式，其他特殊格式都需要单独判断，否则都是按照文本格式统计长度
                lk1=len(str(format(sz,',')))#考虑到数字千分位的情况
            elif sz is None:
                lk1=0
            else:
                lk1=len(str(sz).encode('gbk'))#如果不是gbk格式，两个中文字的长度为2，实际我们需要长度为4
            if lk<lk1:#判断当前单元格长度是不是大于上一个单元格长度，如果大于，则将新的长度赋值到列宽参数上，筛选出长度最大的单元格，并作为列宽数值
                lk=lk1
        lks.append(lk)#存储每一个列宽数值
  #设置列宽
    for i in range(1,ws.max_column+1):#对每一列循环
        k=get_column_letter(i)
        ws.column_dimensions[k].width=min(lks[i-1],20)+2#最大长度为20，避免列宽太大，加2是为了显示内容给人的感觉更宽松一点。
    # 设置行高       
    rows = ws.iter_rows()
    for row in rows:
        for cell in row[:]:
            if cell.value == 'passed':
                cell.font = Font(color = "23B14D")
            elif cell.value == 'broken':
                cell.font = Font(color = "ED1D25") 
            cell.alignment = Alignment(horizontal='left', wrapText=True)  
        # for cell in row[:]:
            cell.alignment = Alignment(horizontal='left', wrapText=True)

wb.save(path)#保存Excel
wb.close()#关闭Excel对象
df_html1 = fwrite.to_html(escape=False,index=False,justify="left")  # DataFrame数据转化为HTML表格形式   
df_html2 = df_html1.replace("<td>passed</td>","<td class='pass'>passed</td>")
df_html = df_html2.replace("<td>broken</td>","<td class='fail'>broken</td>")
#html格式的邮件正文
head = \
        """
        <head>
            <meta charset="utf-8">
            <STYLE TYPE="text/css" MEDIA=screen>
                table.dataframe {
                    border-collapse: collapse;
                    border: 2px solid #a19da2;
                    /*居中显示整个表格*/
                    margin: auto;
                }
                table.dataframe thead {
                    border: 2px solid #91c6e1;
                    background: #f1f1f1;
                    padding: 10px 10px 10px 10px;
                    color: #333333;
                }
                table.dataframe tbody {
                    border: 2px solid #91c6e1;
                    padding: 10px 10px 10px 10px;
                }
                table.dataframe tr {
                }
                table.dataframe th {
                    vertical-align: top;
                    font-size: 14px;
                    padding: 10px 10px 10px 10px;
                    color: #105de3;
                    font-family: arial;
                    text-align: center;
                }
                table.dataframe td {
                    text-align: left;
                    padding: 10px 10px 10px 10px;
                }
                body {
                    font-family: 宋体;
                }
                h1 {
                    color: #5db446
                }
                div.header h2 {
                    color: #0002e3;
                    font-family: 黑体;
                }
                div.content h2 {
                    text-align: center;
                    font-size: 28px;
                    text-shadow: 2px 2px 1px #de4040;
                    color: #fff;
                    font-weight: bold;
                    background-color: #008eb7;
                    line-height: 1.5;
                    margin: 20px 0;
                    box-shadow: 10px 10px 5px #888888;
                    border-radius: 5px;
                }
                h3 {
                    font-size: 22px;
                    background-color: rgba(0, 2, 227, 0.71);
                    text-shadow: 2px 2px 1px #de4040;
                    color: rgba(239, 241, 234, 0.99);
                    line-height: 1.5;
                }
                h4 {
                    color: #e10092;
                    font-family: 楷体;
                    font-size: 20px;
                    text-align: center;
                }
                td img {
                    /*width: 60px;*/
                    max-width: 300px;
                    max-height: 300px;
                }
                .pass {
                    color: #23B14D
                }
                .fail {
                    color: #ED1D25
                }
            </STYLE>
        </head>
        """
body = \
        """
        <body>
        <div align="center" class="header">
            <!--标题部分的信息-->
            <h1 align="center">OpenAI自动化测试结果</h1>
        </div>
        <hr>
        <div class="content">
            <!--正文内容-->
            <h2> </h2>
            </div>
            <div>
                <h4></h4>
                <div align="center">
                    <span style="margin-left:20px">passed：<span class="pass">{pass_num}</span></span>
                    <span style="margin-left:20px">skipped：<span class="skip">{skip_num}</span></span> 
                    <span style="margin-left:20px">broken：<span class="fail">{failed_num}</sapn></span>
                    <p></p>
                </div>
                {df_html}
            </div>
                
            </div>
            <hr>
            <p style="text-align: center">
            </p>
        </div>
        </body>
        """.format(pass_num=pass_num, skip_num=skip_num, failed_num=failed_num, df_html=df_html, time=time_str)
html_msg = "<html>" + head + body + "</html>"
html_msg = html_msg.replace('\n', '').encode("utf-8")

# 邮件服务器设置
smtp_server = 'smtp.feishu.cn'
smtp_port = '465'  # 通常是587或465
username = 'yangshuo@xcoresigma.com'
password = 'PlIncAWZ4s7C11Dk'
 
# 邮件内容设置
sender_email = 'yangshuo@xcoresigma.com'
# receiver_email = ['yangshuo@xcoresigma.com','sunqianqi@xcoresigma.com','zhaojiacheng@xcoresigma.com']
receiver_email = emails
# receiver_email = ['yangshuo@xcoresigma.com','chenlong@xcoresigma.com','zhaojiacheng@xcoresigma.com',
#                   'wangrui@xcoresigma.com','sunqianqi@xcoresigma.com','caowanlu@xcoresigma.com',
#                   'zhangxinyuan@xcoresigma.com','shixiyu@xcoresigma.com','donghanyuan@xcoresigma.com','limingge@xcoresigma.com']

subject = f'xcore_llm api server自动化测试结果'
text_content = '性能测试结果'


# 创建邮件对象
msg = MIMEMultipart()
msg['From'] = sender_email  # 发件人邮箱地址
msg['To'] = ','.join(receiver_email)  # 收件人邮箱地址（飞书邮箱地址）
msg['Subject'] = subject  # 邮件主题
 
# # 添加邮件正文内容
# msg.attach(MIMEText(content, 'html','utf-8'))
msg.attach(MIMEText(html_msg,"html",'utf-8'))

# 构造附件1，txt文件
att1 = MIMEText(open(path, 'rb').read(), 'base64', 'utf-8')
att1["Content-Type"] = 'application/octet-stream'
# 这里的filename可以任意写，写什么名字，邮件中显示什么名字
att1["Content-Disposition"] = f'attachment; filename="{file_name}.xls"'
msg.attach(att1)

smtpObj = smtplib.SMTP_SSL(smtp_server, smtp_port)  # 启用SSL发信, 端口一般是465
smtpObj.login(username,password)
smtpObj.sendmail(sender_email,receiver_email,msg.as_string())
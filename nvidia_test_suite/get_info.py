import argparse
import os
import re
import smtplib
from collections import OrderedDict
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import pandas as pd
import pytz
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


parser = argparse.ArgumentParser(description="解析 UnitTest 日志并生成报告")
parser.add_argument("--file", type=str, required=True, help="待解析日志路径")
parser.add_argument("--email", type=str, default="")
parser.add_argument("--env", type=str, default="")
parser.add_argument("--url", type=str, default="")
parser.add_argument("--model", type=str, default="")
parser.add_argument("--gpu", type=str, default="")
parser.add_argument("--cmd", type=str, default="")
args = parser.parse_args()

log_path = args.file
emails = [x.strip() for x in args.email.split(",") if x.strip()]
model = args.model
gpu = args.gpu
cmd = args.cmd

if not os.path.exists(log_path):
    raise FileNotFoundError(f"日志文件不存在: {log_path}")

tz = pytz.timezone("Asia/Shanghai")
time_str = datetime.now(tz).strftime("%Y_%m_%d")
log_stem = os.path.splitext(os.path.basename(log_path))[0]
excel_path = f"./{log_stem}.xlsx"
html_path = f"./{log_stem}.html"


def parse_unittest_log(file_path):
    """
    解析 pytest 输出日志，提取每个用例最终状态。
    兼容示例:
    [gw1] [ 12%] PASSED tests/test_xxx.py::test_a[param]
    """
    case_status_map = OrderedDict()
    case_pattern = re.compile(r"\b(PASSED|FAILED|SKIPPED|XFAIL|XPASS|ERROR)\b\s+(tests/\S+)")
    summary_pattern = re.compile(
        r"(?:(\d+)\s+passed)?(?:,\s*)?(?:(\d+)\s+skipped)?(?:,\s*)?(?:(\d+)\s+failed)?",
        re.IGNORECASE,
    )

    summary = {"passed": 0, "skipped": 0, "failed": 0}
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        for raw_line in f:
            line = raw_line.strip()
            case_match = case_pattern.search(line)
            if case_match:
                status = case_match.group(1).lower()
                case_name = case_match.group(2)
                normalized = "failed" if status in ("failed", "error", "xpass") else ("skipped" if status in ("skipped", "xfail") else "passed")
                case_status_map[case_name] = normalized
                continue

            if "passed in" in line and ("passed" in line or "failed" in line or "skipped" in line):
                summary_match = summary_pattern.search(line)
                if summary_match:
                    p, s, f_num = summary_match.groups()
                    summary["passed"] = int(p) if p else 0
                    summary["skipped"] = int(s) if s else 0
                    summary["failed"] = int(f_num) if f_num else 0

    if case_status_map:
        counts = {"passed": 0, "skipped": 0, "failed": 0}
        for status in case_status_map.values():
            counts[status] += 1
        summary = counts

    records = [{"用例名称": name, "运行结果": status} for name, status in sorted(case_status_map.items(), key=lambda x: x[0])]
    return records, summary


records, summary = parse_unittest_log(log_path)
fwrite = pd.DataFrame(records if records else [{"用例名称": "未解析到用例", "运行结果": "failed"}])
fwrite.to_excel(excel_path, index=False, engine="openpyxl")

wb = load_workbook(excel_path)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    column_widths = []
    for i in range(1, ws.max_column + 1):
        width = 1
        for j in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=j, column=i).value
            if isinstance(cell_value, (int, float)):
                candidate = len(str(format(cell_value, ",")))
            elif cell_value is None:
                candidate = 0
            else:
                candidate = len(str(cell_value).encode("gbk", errors="ignore"))
            width = max(width, candidate)
        column_widths.append(width)

    for i in range(1, ws.max_column + 1):
        k = get_column_letter(i)
        ws.column_dimensions[k].width = min(column_widths[i - 1], 20) + 2
    # 第二列“运行结果”固定更宽，提升可读性
    ws.column_dimensions[get_column_letter(2)].width = max(ws.column_dimensions[get_column_letter(2)].width, 16)

    for row in ws.iter_rows():
        for cell in row:
            if cell.value == "passed":
                cell.font = Font(color="23B14D")
            elif cell.value == "failed":
                cell.font = Font(color="ED1D25")
            elif cell.value == "skipped":
                cell.font = Font(color="999999")
            cell.alignment = Alignment(horizontal="left", wrapText=True)

wb.save(excel_path)
wb.close()

fwrite_html = fwrite.copy()
# 防止邮箱客户端将 `*.py::test_xxx` 自动识别为链接
fwrite_html["用例名称"] = fwrite_html["用例名称"].str.replace(".py::", ".py&#8203;::", regex=False)

df_html = fwrite_html.to_html(escape=False, index=False, justify="left")
df_html = df_html.replace("<td>passed</td>", "<td class='pass'>passed</td>")
df_html = df_html.replace("<td>failed</td>", "<td class='fail'>failed</td>")
df_html = df_html.replace("<td>skipped</td>", "<td class='skip'>skipped</td>")

head = """
<head>
    <meta charset="utf-8">
    <style type="text/css" media="screen">
        table.dataframe { border-collapse: collapse; border: 2px solid #a19da2; margin: auto; }
        table.dataframe thead { border: 2px solid #91c6e1; background: #f1f1f1; color: #333333; }
        table.dataframe tbody { border: 2px solid #91c6e1; }
        table.dataframe th { vertical-align: top; font-size: 14px; padding: 10px; color: #105de3; text-align: center; }
        table.dataframe td { text-align: left; padding: 10px; }
        table.dataframe th:nth-child(2), table.dataframe td:nth-child(2) { min-width: 120px; text-align: center; }
        body { font-family: sans-serif; }
        .pass { color: #23B14D; }
        .fail { color: #ED1D25; }
        .skip { color: #999999; }
    </style>
</head>
"""

body = """
<body>
    <div align="left" class="header">
        <div>显卡类型：{gpu}</div>
        <div>测试模型：{model}</div>
        <div>启动命令：{cmd}</div>
        <div>日志文件：{log_name}</div>
        <div>生成日期：{time_str}</div>
    </div>
    <hr>
    <div align="center">
        <span style="margin-left:20px">passed：<span class="pass">{pass_num}</span></span>
        <span style="margin-left:20px">skipped：<span class="skip">{skip_num}</span></span>
        <span style="margin-left:20px">failed：<span class="fail">{failed_num}</span></span>
    </div>
    <p></p>
    {df_html}
    <hr>
</body>
""".format(
    gpu=gpu,
    model=model,
    cmd=cmd,
    log_name=os.path.basename(log_path),
    time_str=time_str,
    pass_num=summary["passed"],
    skip_num=summary["skipped"],
    failed_num=summary["failed"],
    df_html=df_html,
)

html_msg = ("<html>" + head + body + "</html>").replace("\n", "")
with open(html_path, "w", encoding="utf-8") as f:
    f.write(html_msg)

# 邮件服务器设置
smtp_server = "smtp.feishu.cn"
smtp_port = "465"
username = "yangshuo@xcoresigma.com"
password = "PlIncAWZ4s7C11Dk"
sender_email = "limingge@xcoresigma.com"

if emails:
    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = ",".join(emails)
    msg["Subject"] = "xcore_llm api server自动化测试结果"
    msg.attach(MIMEText(html_msg, "html", "utf-8"))

    with open(excel_path, "rb") as fp:
        att_excel = MIMEText(fp.read(), "base64", "utf-8")
    att_excel["Content-Type"] = "application/octet-stream"
    att_excel["Content-Disposition"] = f'attachment; filename="{os.path.basename(excel_path)}"'
    msg.attach(att_excel)

    if os.path.exists("./test_info.log"):
        with open("./test_info.log", "rb") as fp:
            att_log = MIMEText(fp.read(), "base64", "utf-8")
        att_log["Content-Type"] = "application/octet-stream"
        att_log["Content-Disposition"] = 'attachment; filename="test_info.log"'
        msg.attach(att_log)

    smtp_obj = smtplib.SMTP_SSL(smtp_server, smtp_port)
    smtp_obj.login(username, password)
    smtp_obj.sendmail(sender_email, emails, msg.as_string())
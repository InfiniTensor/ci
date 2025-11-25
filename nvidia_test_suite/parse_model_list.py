#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
解析model_list.xlsx文件，提取模型名称、卡型和tp参数
"""

import re
from openpyxl import load_workbook

def extract_tp_number(default_params):
    """从默认参数中提取-tp <num>中的num"""
    if not default_params:
        return None
    
    # 匹配 -tp 后面的数字
    pattern = r'-tp\s+(\d+)'
    match = re.search(pattern, str(default_params))
    if match:
        return match.group(1)
    else:
        pattern = r'--tensor-parallel-size\s+(\d+)'
        match = re.search(pattern, str(default_params))
        if match:
            return match.group(1)
    return None

def clean_card_type(card_type):
    """清理卡型，去掉括号中的内容和额外描述"""
    if not card_type:
        return None
    
    card_type = str(card_type).strip()
    # 去掉括号及其内容，例如 'H20 (96G)' -> 'H20'
    card_type = re.sub(r'\s*\([^)]*\)', '', card_type)
    # 去掉额外的描述文本（如"长上下文"等）
    # 只保留卡型名称部分（通常是字母数字组合，可能包含斜杠）
    card_type = re.sub(r'\s+[^\w/]+.*$', '', card_type)
    return card_type.strip()

def extract_card_types(card_type_str):
    """从卡型字符串中提取所有卡型（可能包含多个，用顿号分隔）"""
    if not card_type_str:
        return []
    
    card_type_str = str(card_type_str).strip()
    # 按顿号分割
    card_types = [ct.strip() for ct in card_type_str.split('、')]
    
    # 清理每个卡型并过滤
    cleaned_cards = []
    for ct in card_types:
        cleaned = clean_card_type(ct)
        if cleaned:
            # 如果卡型包含斜杠（如H100/A100），需要分别处理
            if '/' in cleaned:
                parts = [p.strip() for p in cleaned.split('/')]
                for part in parts:
                    cleaned_cards.append(part)
            else:
                cleaned_cards.append(cleaned)
    
    return cleaned_cards

def parse_model_list(card_type, excel_path):
    """解析Excel文件并提取指定卡型的模型列表"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb.active
    
    results = []
    
    # 从第二行开始读取（第一行是表头）
    for row in ws.iter_rows(min_row=2, values_only=True):
        model_name = row[0]
        card_type_list = row[1]
        default_params = row[2]
        
        # 跳过空行
        if not model_name:
            continue
        
        # 提取所有有效的卡型
        card_types = extract_card_types(card_type_list)
        
        # 如果没有有效卡型，跳过
        if not card_types:
            continue
        
        # 提取tp参数
        tp_num = extract_tp_number(default_params)
        
        if tp_num:
            # 为每个卡型生成一条记录
            for clean_card in card_types:
                if clean_card == card_type:
                    # 合并成格式：模型名称:<num>:卡型
                    result = f"{model_name}:{tp_num}:{clean_card}"
                    results.append(result)
    
    return results

def main():
    import sys  
    if len(sys.argv) < 3:
        print("Usage: python parse_model_list.py <card_type> <excel_path>")
        sys.exit(1)
    
    card_type = sys.argv[1]
    excel_path = sys.argv[2]
    results = parse_model_list(card_type, excel_path)
    
    # 输出结果
    for result in results:
        print(result)

if __name__ == '__main__':
    main()

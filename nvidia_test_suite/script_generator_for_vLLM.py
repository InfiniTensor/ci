from openpyxl import load_workbook
import re
import os
import sys

def clean_card_type(card_type):
    """清理卡型，去掉括号中的内容和额外描述"""
    if not card_type:
        return None
    
    card_type = str(card_type).strip()
    # 去掉括号及其内容，例如 'H20 (96G)' -> 'H20'
    card_type = re.sub(r'\s*\([^)]*\)', '', card_type)
    # 去掉额外的描述文本（如"长上下文"等）
    # 只保留卡型名称部分（通常是字母数字组合，可能包含斜杠）
    card_type = re.sub(r'\s+[^\w/]+.*$', '', card_type)
    return card_type.strip()

def extract_card_types(card_type_str):
    """从卡型字符串中提取所有卡型（可能包含多个，用顿号分隔）"""
    if not card_type_str:
        return []
    
    card_type_str = str(card_type_str).strip()
    # 按顿号分割
    card_types = [ct.strip() for ct in card_type_str.split('、')]
    
    # 清理每个卡型并过滤
    cleaned_cards = []
    for ct in card_types:
        cleaned = clean_card_type(ct)
        if cleaned:
            # 如果卡型包含斜杠（如H100/A100），需要分别处理
            if '/' in cleaned:
                parts = [p.strip() for p in cleaned.split('/')]
                for part in parts:
                    cleaned_cards.append(part)
            else:
                cleaned_cards.append(cleaned)
    
    return cleaned_cards

def main():
    verison = ""
    if len(sys.argv) != 3:
        print("Usage: python script_generator <test_type> <version>")
        sys.exit(1)
    
    test_type = sys.argv[1]
    version = sys.argv[2]
    
    curr_dir = os.getcwd()

    # 加载 Excel 文件
    file_path = f'{curr_dir}/{version}/model_list.xlsx'  # 替换为你的 Excel 文件路径
    workbook = load_workbook(file_path)

    # 选择工作表
    sheet = workbook['NVIDIA']

    # 获取行数
    row_count = sheet.max_row
    print(f"总行数: {row_count}")
    
    target_file = ""
    src_code = ""
    
    if test_type == "Smoke":
        target_file = "vLLM_job_executor_for_SmokeTest.sh"
        src_code += "docker exec vllm_nvidia_SmokeTest_${JOB_COUNT} /bin/bash -c \"\n"
    elif test_type == "Performance":
        target_file = "vLLM_job_executor_for_PerformanceTest.sh"
        src_code += "docker exec vllm_nvidia_PerformanceTest_${JOB_COUNT} /bin/bash -c \"\n"
    elif test_type == "Stability":
        target_file = "vLLM_job_executor_for_StabilityTest.sh"
        src_code += "docker exec vllm_nvidia_StabilityTest_${JOB_COUNT} /bin/bash -c \"\n"
    elif test_type == "Accuracy":
        target_file = "vLLM_job_executor_for_AccuracyTest.sh"
        src_code += "docker exec vllm_nvidia_AccuracyTest_${JOB_COUNT} /bin/bash -c \"\n"
    
    start = True
    for row in sheet.iter_rows(min_row=2, max_row=row_count, values_only=True):
        # print(row)  # 每行数据以元组形式返回
        name = row[0]
        GPU = row[1]
        args = row[2]
        args = args.split('\n')[0]

        result = re.sub(r"^.*docker\.xcoresigma\.com/docker/vllm/vllm-openai\:\S+", "", args)
        result = re.sub(r"--model\s+", "", result)
        result = re.sub(r"--port\s+\d+", "--port $PORT", result)
        result = re.sub(r"--served-model-name\s+\S+", f"--served-model-name {name}", result)
        
        card_types = extract_card_types(GPU)
        for card_type in card_types:
            if start:
                src_code += f"if [ \\\"$MODEL_$GPU_MODEL\\\" == \\\"{name}_{card_type}\\\" ]; then\n"
                start = False
            else:
                src_code += f"elif [ \\\"$MODEL_$GPU_MODEL\\\" == \\\"{name}_{card_type}\\\" ]; then\n"
            src_code += "    echo \\\"vllm serve "
            src_code += result
            src_code += "\\\"\n"
            src_code += "    nohup env CUDA_DEVICE_ORDER=PCI_BUS_ID CUDA_VISIBLE_DEVICES=$CUDA_VISIBLE_DEVICES vllm serve "
            src_code += result
            src_code += " > $LOG_NAME 2>&1 &\n"
    src_code += "fi\n"

    # print(src_code)

    template_file = "job_executor_template_for_vLLM.sh"

    try:
        with open(f"{curr_dir}/{template_file}", 'r', encoding='utf-8') as file:
            lines = file.readlines()
        
        line_num = 0
        for line in lines:
            if "<<<generated source code>>>" in line:
                lines[line_num] = src_code
                with open(f"{curr_dir}/{target_file}", 'w') as file:
                    file.write(''.join(lines))
                # print(lines)
                break
            elif "<<<TEST_TYPE>>>" in line:
                if test_type == "Smoke":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "SmokeTest")
                elif test_type == "Performance":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "PerformanceTest")
                elif test_type == "Accuracy":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "AccuracyTest")
                elif test_type == "Stability":
                    lines[line_num] = line.replace("<<<TEST_TYPE>>>", "StabilityTest")
            line_num += 1
    except FileNotFoundError:
        print(f"Error: Log file '{curr_dir}/{template_file}' not found.")
    except Exception as e:
        print(f"Error reading file: {str(e)}")
    
    os.system(f"chmod 777 {target_file}")

if __name__ == "__main__":
    main()
